import openpyxlimport shutilimport datetimeimport osimport refrom openpyxl.styles import Color, PatternFill, Font, Borderfrom openpyxl.styles import colorsfrom openpyxl.cell import CellredFill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')def Prod_ID_Block_baseline(s, n):                for start in range(24, len(s), n):                                         yield s[start:start+n]def Prod_ID_Block_New(s, n):                for start in range(24, len(s), n):                                         yield s[start:start+n]Lines_Base = []Lines_New = []def Find_Occurrence_PCID(File_Name, PCID,Lines):        Inc1 =0        Inc2=0        del Lines[:]     with open(File_Name, 'r') as f2:                for line in f2:                        if Inc2 not in {0, num_lines-1}:                                Inc2+=1                                for Prod_ID_Block in Prod_ID_Block_baseline(line, 240):                                        if Prod_ID_Block.strip() not in ("","00000000000"):                                                if len(Prod_ID_Block.strip())>15:                            keyword1 = re.compile(r'%s'%PCID)                                                        test = keyword1.search(Prod_ID_Block)                            if test != None:                                #if Product_Cust_ID == 'CFP33K ' :                                    #print(Lines_Base[Baseline_Row])                                    #print(Prod_ID_Block)                                                                Lines_Base.append(Prod_ID_Block)                                                                Inc1+=1                                            else:                                Inc2+=1                    return Inc1def Find_Occurrence_PCID_New(File_Name, PCID,Lines1):        Inc1 =0        Inc2=0        #del Lines1[:]     with open(File_Name, 'r') as f2:                for line in f2:                        if Inc2 not in {0, num_lines-1}:                                Inc2+=1                                for Prod_ID_Block_Neww in Prod_ID_Block_New(line, 240):                                        if Prod_ID_Block_Neww.strip() not in ("","00000000000"):                                                if len(Prod_ID_Block_Neww.strip())>15:                            keyword1 = re.compile(r'%s'%PCID)                                                        test = keyword1.search(Prod_ID_Block_Neww)                            if test != None:                                #if Product_Cust_ID == 'CFP33K ' :                                    #print(Lines_Base[Baseline_Row])                                    #print(Prod_ID_Block)                                                                Lines_New.append(Prod_ID_Block_Neww)                                                                Inc1+=1                                            else:                                Inc2+=1                    return Inc1                redFill = PatternFill(fill_type='solid', start_color='F2DCDB',end_color='F2DCDB')# Funtion to indentify position of Prod Customer IDdef find_str(s, char):    index = 0    if char in s:        c = char[0]        for ch in s:            if ch == c:                if s[index:index+len(char)] == char:                    return index            index += 1    return -1##### Creating Product level tabs for results #########$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$now = datetime.datetime.now()os.mkdir('C:/PosterFile_Compare/MasterFile/')Master_workbook = openpyxl.Workbook()Master_workbook.create_sheet("Compare")Master_Compare_sheet = Master_workbook ['Compare']TC_Row1 = 1    for file_Base in os.listdir("C:/PosterFile_Compare/Baseline"):        if file_Base.endswith(".dat"):        for file_New in os.listdir("C:/PosterFile_Compare/FilesToCompare"):        #print("Test")                    Test1 = os.path.join("C:/PosterFile_Compare/Baseline", file_Base)                        Test2 = os.path.join("C:/PosterFile_Compare/FilesToCompare", file_New)            if Test1[len(Test1) - 20: len(Test1)-13] == Test2[len(Test2) - 20: len(Test2)-13]:                            Master_Compare_sheet.cell(row =TC_Row1,column = 1).value = Test1[len(Test1) - 20: len(Test1)-13]                            Master_Compare_sheet.cell(row =TC_Row1,column = 2).value = Test1                            Master_Compare_sheet.cell(row =TC_Row1,column = 3).value = Test2                            TC_Row1 = TC_Row1 +1Master_workbook.save("C:/PosterFile_Compare/MasterFile/Master_File.xlsx")Validation_Rows1 = Master_Compare_sheet.max_rowTC_Row2 = 1Folder_time = now.strftime("%Y-%m-%d-%H-%M-%S")path = os.mkdir('C:/PosterFile_Compare/Logs/'+ now.strftime("%Y-%m-%d-%H-%M-%S"))for No_Files in range(0, Validation_Rows1):        Product_Name = Master_Compare_sheet.cell(row =TC_Row2,column = 1).value     Baseline_File = Master_Compare_sheet.cell(row =TC_Row2,column = 2).value    File_To_Compare = Master_Compare_sheet.cell(row =TC_Row2,column = 3).value        #$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$    List_Multiple_Not_Matching_Baseline = list()    List_Multiple_Not_Matching_New = list()    List_Block_Baseline= list()    List_Block_New= list()    List_Block_Baseline1 = list()    List_Block_New1 = list()    Master_File = Baseline_File    File_To_Compare = File_To_Compare        Baseline_File = open(Master_File, 'r')    FileCompare = open(File_To_Compare, 'r')    # Make each line in a list for baseline and new files    list1=Baseline_File.readlines()    list2=FileCompare.readlines()    # Get no of lines for baseline and new files    Baseline_No_Rows = len(list1)    FileToCompare_No_Rows = len(list2)     # Split Line of data into 6 blocks for baseline files    for r in range(1, Baseline_No_Rows-1):        Baseline_Value = list1[r]           for Prod_ID_Block in Prod_ID_Block_baseline(Baseline_Value, 240):            #print (Prod_ID_Block)            List_Block_Baseline.append(Prod_ID_Block)    # Split Line of data into 6 blocks for new files    for r in range(1, FileToCompare_No_Rows-1):        New_Value = list2[r]            for Prod_ID_Block_Newww in Prod_ID_Block_New(New_Value, 240):                        List_Block_New.append(Prod_ID_Block_Newww)    List_Block_Baseline = list(filter(None, List_Block_Baseline))    List_Block_New = list(filter(None, List_Block_New))    Baseline_No_PCID = len(List_Block_Baseline)    FileToCompare_No_PCID =  len(List_Block_New)    for i in range(0,Baseline_No_PCID):        #print (List_Block_Baseline[i])                test = List_Block_Baseline[i]        Acc = test[22:37].strip()        PCID = test[153:166].strip()                if (PCID != ""):            print (PCID)            List_Block_Baseline1.append(PCID)    for i in range(0,FileToCompare_No_PCID):        #print (List_Block_New[i])        test = List_Block_New[i]        Acc = test[22:37].strip()        PCID = test[153:166].strip()        if (PCID != ""):            print (PCID)            List_Block_New1.append(PCID)    Master_List = list(set(List_Block_Baseline1+List_Block_New1))    No_PCID = len(Master_List)       Master_workbook.create_sheet(Product_Name)    Master_Compare_sheet_Prod = Master_workbook [Product_Name]    for i in range(1,No_PCID+1):                       Master_Compare_sheet_Prod.cell(row =i+1,column = 1).value = Master_List[i-1]    Master_Compare_sheet_Prod.cell(row =1,column = 1).value = "Product_Cust_ID"    Master_Compare_sheet_Prod.cell(row =1,column = 2).value = "Matching/Not-Matching"    #Master_workbook.save('C:/Test1/Logs/'+ Folder_time + '/' "Master_File_Results.xlsx")    Master_workbook.save("C:/PosterFile_Compare/Logs/" + now.strftime("%Y-%m-%d-%H-%M-%S") + "/Master_File_Results"+now.strftime("%Y-%m-%d-%H-%M-%S")+".xlsx")    TC_Row2 = TC_Row2 + 1shutil.rmtree('C:/PosterFile_Compare/MasterFile/')#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%##### Main script start here #######Master_Compare_sheet = Master_workbook ['Compare']##### Identify no of rows to validate#######Validation_Rows = Master_Compare_sheet.max_rowValidation_Row = 1status = "False"while (Validation_Rows !=0):    Product = Master_Compare_sheet.cell(row =Validation_Row,column = 1).value    BaseLine_File = Master_Compare_sheet.cell(row =Validation_Row,column = 2).value    File_To_Compare = Master_Compare_sheet.cell(row =Validation_Row,column = 3).value    Validation_Sheet = Master_workbook [Product]        No_Of_PCID = (Validation_Sheet.max_row)        for TC_Row in range(2,No_Of_PCID+1):        #print(Product)        match_count = 0        Product_Cust_ID = str(Validation_Sheet.cell(row =TC_Row,column = 1).value) + " "        num_lines = sum(1 for line in open(BaseLine_File))           # Calling function to identify row number of Prod Customer ID                Baseline_PCID_Count = Find_Occurrence_PCID(BaseLine_File, Product_Cust_ID, Lines_Base)        num_lines = sum(1 for line in open(File_To_Compare))                FilesToCompare_PCID_Count = Find_Occurrence_PCID_New(File_To_Compare, Product_Cust_ID, Lines_New)                    if (Baseline_PCID_Count < 2 and FilesToCompare_PCID_Count < 2):                                if (Baseline_PCID_Count != 0 and FilesToCompare_PCID_Count != 0 ):                                                               if (Lines_Base[0] != Lines_New[0]):                                                                        Validation_Sheet.cell(row =TC_Row,column = 2).value = "Not Matching"                        Validation_Sheet.cell(row =TC_Row,column = 2).fill = redFill                                                Baseline_Value_Issue = Lines_Base[0]                                                New_Value_issue = Lines_New[0]                        Validation_Sheet.cell(row =TC_Row,column = 3).value = Baseline_Value_Issue                        Validation_Sheet.cell(row =TC_Row,column = 4).value = New_Value_issue                                   Master_workbook.save("C:/PosterFile_Compare/Logs/" + now.strftime("%Y-%m-%d-%H-%M-%S") + "/Master_File_Results"+now.strftime("%Y-%m-%d-%H-%M-%S")+".xlsx")                                          TC_Row = TC_Row + 1                    else:                                                Validation_Sheet.cell(row =TC_Row,column = 2).value = "Matching"                        Master_workbook.save("C:/PosterFile_Compare/Logs/" + now.strftime("%Y-%m-%d-%H-%M-%S") + "/Master_File_Results"+now.strftime("%Y-%m-%d-%H-%M-%S")+".xlsx")                                                TC_Row = TC_Row + 1                                        elif (Baseline_PCID_Count == 0 and FilesToCompare_PCID_Count != 0 ):                    Validation_Sheet.cell(row =TC_Row,column = 2).value = "PCID not found in baseline file"                    Validation_Sheet.cell(row =TC_Row,column = 2).fill = redFill                    Master_workbook.save("C:/PosterFile_Compare/Logs/" + now.strftime("%Y-%m-%d-%H-%M-%S") + "/Master_File_Results"+now.strftime("%Y-%m-%d-%H-%M-%S")+".xlsx")                                        TC_Row = TC_Row + 1                elif (Baseline_PCID_Count != 0 and FilesToCompare_PCID_Count == 0 ):                    Validation_Sheet.cell(row =TC_Row,column = 2).value = "PCID not found in new file"                    Validation_Sheet.cell(row =TC_Row,column = 2).fill = redFill                    Master_workbook.save("C:/PosterFile_Compare/Logs/" + now.strftime("%Y-%m-%d-%H-%M-%S") + "/Master_File_Results"+now.strftime("%Y-%m-%d-%H-%M-%S")+".xlsx")                                        TC_Row = TC_Row + 1                else:                    Validation_Sheet.cell(row =TC_Row,column = 2).value = "PCID not found in both files"                    Validation_Sheet.cell(row =TC_Row,column = 2).fill = redFill                    Master_workbook.save('C:/PosterFile_Compare/Logs/'+ now.strftime("%Y-%m-%d-%H-%M-%S")+ '/Master_File_Results.xlsx')                    TC_Row = TC_Row + 1        elif (Baseline_PCID_Count == FilesToCompare_PCID_Count):            #print(Baseline_PCID_Count, FilesToCompare_PCID_Count)            Inc = 0            match_count = 0           # print (len(Lines_New),len(Lines_Base))           # print (Product)            Lines_New = list(filter(None, Lines_New))            Lines_Base = list(filter(None, Lines_Base))            #print (len(Lines_New),len(Lines_Base))                        #print(Lines_New[Newfile_Row])                                      for Baseline_Row in range(0, Baseline_PCID_Count):                                                status = "False"                                Acc_Found = "True"                                for Newfile_Row in range(0, FilesToCompare_PCID_Count):                    #print(Lines_New[Newfile_Row])                                                                              if Lines_Base[Baseline_Row].strip() == Lines_New[Newfile_Row].strip():                                                status = "True"                        match_count = match_count +1                                if status == "True":                    status = "True"                    #match_count = match_count +1                else:                    List_Multiple_Not_Matching_Baseline.append(Lines_Base[Baseline_Row])                    #print(List_Multiple_Not_Matching_Baseline[0])                            if match_count == Baseline_PCID_Count:                                                Validation_Sheet.cell(row =TC_Row,column = 2).value = "Multiple Entries- Matching"                Master_workbook.save("C:/PosterFile_Compare/Logs/" + now.strftime("%Y-%m-%d-%H-%M-%S") + "/Master_File_Results"+now.strftime("%Y-%m-%d-%H-%M-%S")+".xlsx")            else:                Validation_Sheet.cell(row =TC_Row,column = 2).value = "Multiple Entries- Not Matching"                Validation_Sheet.cell(row =TC_Row,column = 2).fill = redFill                Master_workbook.save("C:/PosterFile_Compare/Logs/" + now.strftime("%Y-%m-%d-%H-%M-%S") + "/Master_File_Results"+now.strftime("%Y-%m-%d-%H-%M-%S")+".xlsx")                                NM_Count = len(List_Multiple_Not_Matching_Baseline)                #print (NM_Count)                for NM_Count_col in range(0,NM_Count):                    Validation_Sheet.cell(row =TC_Row,column = NM_Count_col + 3).value = List_Multiple_Not_Matching_Baseline[NM_Count_col]                                                                               TC_Row = TC_Row + 1        #status == "True":        del List_Multiple_Not_Matching_Baseline [:]        del Lines_New[:]            del Lines_Base[:]        del List_Block_Baseline[:]        del List_Block_New[:]        del List_Block_New1[:]        del List_Block_Baseline1[:]    Validation_Rows = Validation_Rows -1    Validation_Row = Validation_Row +1                print ( "Poster files compare completed. Please check Logs folder for results!")