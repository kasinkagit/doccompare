import openpyxlimport shutilimport datetimeimport osimport refrom openpyxl.styles import Color, PatternFill, Font, Borderfrom openpyxl.styles import colorsfrom openpyxl.cell import CellredFill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')def Prod_ID_Block_baseline(s, n):                for start in range(25, len(s), n):                                         yield s[start:start+n]Lines_Base = []Lines_New = []def Find_Occurrence_PCID(File_Name, PCID,Lines):        Inc1 =0        Inc2=0        with open(File_Name, 'r') as f2:                for line in f2:                        if Inc2 not in {0, num_lines-1}:                                Inc2+=1                                for Prod_ID_Block in Prod_ID_Block_baseline(line, 240):                                        if Prod_ID_Block.strip() not in ("","00000000000"):                                                if len(Prod_ID_Block.strip())>15:                            keyword1 = re.compile(r'%s'%PCID)                                                        test = keyword1.search(Prod_ID_Block)                            if test != None:                                                                Lines.append(Prod_ID_Block)                                                                Inc1+=1                                            else:                                Inc2+=1                    return Inc1                redFill = PatternFill(fill_type='solid', start_color='F2DCDB',end_color='F2DCDB')# Funtion to indentify position of Prod Customer IDdef find_str(s, char):    index = 0    if char in s:        c = char[0]        for ch in s:            if ch == c:                if s[index:index+len(char)] == char:                    return index            index += 1    return -1##### Main script start here ########$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$now = datetime.datetime.now()os.mkdir('C:/Test1/MasterFile/')Master_workbook = openpyxl.Workbook()Master_workbook.create_sheet("Compare")Master_Compare_sheet = Master_workbook ['Compare']TC_Row1 = 1    for file in os.listdir("C:/Test1/Baseline"):    if file.endswith(".dat"):        Test1 = os.path.join("C:/Test1/Baseline", file)        Test2 = os.path.join("C:/Test1/FilesToCompare", file)        Master_Compare_sheet.cell(row =TC_Row1,column = 1).value = Test1[len(Test1) - 20: len(Test1)-13]        Master_Compare_sheet.cell(row =TC_Row1,column = 2).value = Test1        Master_Compare_sheet.cell(row =TC_Row1,column = 3).value = Test2            TC_Row1 = TC_Row1 +1Master_workbook.save("C:/Test1/MasterFile/Master_File.xlsx")Validation_Rows = Master_Compare_sheet.max_rowTC_Row = 2Folder_time = now.strftime("%Y-%m-%d-%H-%M-%S")path = os.mkdir('C:/Test1/Logs/'+ now.strftime("%Y-%m-%d-%H-%M-%S"))for No_Files in range(2, Validation_Rows-3):    Product_Name = Master_Compare_sheet.cell(row =TC_Row,column = 1).value     Baseline_File = Master_Compare_sheet.cell(row =TC_Row,column = 2).value    File_To_Compare = Master_Compare_sheet.cell(row =TC_Row,column = 3).value        #$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$    List_Block_Baseline= list()    List_Block_New= list()    List_Block_Baseline1 = list()    List_Block_New1 = list()    Master_File = Baseline_File    File_To_Compare = Baseline_File    Baseline_File = open(Master_File, 'r')    FileCompare = open(File_To_Compare, 'r')    #FO = open('C:/Test1/Logs/Log11.txt', 'w')    # Make each line in a list for baseline and new files    list1=Baseline_File.readlines()    list2=FileCompare.readlines()    # Get no of lines for baseline and new files    Baseline_No_Rows = len(list1)    FileToCompare_No_Rows = len(list2)     # Split Line of data into 6 blocks for baseline files    for r in range(1, Baseline_No_Rows-1):        Baseline_Value = list1[r]           for Prod_ID_Block in Prod_ID_Block_baseline(Baseline_Value, 240):            List_Block_Baseline.append(Prod_ID_Block)    # Split Line of data into 6 blocks for new files    for r in range(1, FileToCompare_No_Rows-1):        New_Value = list2[r]            for Prod_ID_Block in Prod_ID_Block_baseline(New_Value, 240):            List_Block_New.append(Prod_ID_Block)    List_Block_Baseline = list(filter(None, List_Block_Baseline))    List_Block_New = list(filter(None, List_Block_New))    Baseline_No_PCID = len(List_Block_Baseline)    FileToCompare_No_PCID =  len(List_Block_New)        for i in range(0,Baseline_No_PCID):        test = List_Block_Baseline[i]        Acc = test[22:37].strip()        PCID = test[152:162].strip()        if (PCID != ""):            List_Block_Baseline1.append(PCID)    for i in range(0,FileToCompare_No_PCID):        test = List_Block_New[i]        Acc = test[22:37].strip()        PCID = test[152:162].strip()        if (PCID != ""):            List_Block_New1.append(PCID)    Master_List = list(set(List_Block_Baseline1+List_Block_New1))    No_PCID = len(Master_List)       #Master_File_Loc = "C:/Test1/MasterFile/Master_File1"+ Product_Name + ".xlsx"    #Master_workbook = openpyxl.load_workbook(Master_File_Loc,read_only = False)    #Master_workbook1 = openpyxl.Workbook()    Master_workbook.create_sheet(Product_Name)    Master_Compare_sheet_Prod = Master_workbook [Product_Name]    for i in range(2,No_PCID+2):        Master_Compare_sheet_Prod.cell(row =i+1,column = 1).value = Master_List[i-2]    Master_workbook.save('C:/Test1/Logs/'+ Folder_time + '/' "Master_File_Results.xlsx")    #Master_workbook.save   # Master_workbook1.save(Master_File_Loc)    TC_Row = TC_Row + 1shutil.rmtree('C:/Test1/MasterFile/')#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%##### Main script start here ########Master_File_Loc = "C:/Test1/MasterFile/Master_File.xlsx"#Master_workbook = openpyxl.load_workbook(Master_File_Loc,read_only = False)Master_Compare_sheet = Master_workbook ['Compare']##### Identify no of rows to validate#######Validation_Rows = Master_Compare_sheet.max_rowprint (Validation_Rows)TC_Row = 2status = "False"while (Validation_Rows !=1):    Product_Cust_ID = str(Master_Compare_sheet.cell(row =TC_Row,column = 1).value) + " "    Product = Master_Compare_sheet.cell(row =TC_Row,column = 2).value    BaseLine_File = Master_Compare_sheet.cell(row =TC_Row,column = 3).value    File_To_Compare = Master_Compare_sheet.cell(row =TC_Row,column = 4).value    # Calling function to identify row number of Prod Customer ID    Baseline_PCID_Count = Find_Occurrence_PCID(BaseLine_File, Product_Cust_ID, Lines_Base)        FilesToCompare_PCID_Count = Find_Occurrence_PCID(File_To_Compare, Product_Cust_ID, Lines_New)        if (Baseline_PCID_Count < 2 and FilesToCompare_PCID_Count < 2):            if (Baseline_PCID_Count != 0 and FilesToCompare_PCID_Count != 0 ):                                FO = open ('C:/Test1/Logs/'+ now.strftime("%Y-%m-%d-%H-%M-%S")+ '/log.txt', "a")                                if (Lines_Base[0] != Lines_New[0]):                                        Master_Compare_sheet.cell(row =TC_Row,column = 5).value = "Not Matching"                    Master_Compare_sheet.cell(row =TC_Row,column = 5).fill = redFill                                        Baseline_Value_Issue = Lines_Base[0]                                        New_Value_issue = Lines_New[0]                    Master_Compare_sheet.cell(row =TC_Row,column = 6).value = Baseline_Value_Issue                    Master_Compare_sheet.cell(row =TC_Row,column = 7).value = New_Value_issue                               Master_workbook.save('C:/Test1/Logs/'+ now.strftime("%Y-%m-%d-%H-%M-%S")+ '/Master_File_Results.xlsx')                    Log_File.writelines("\n[" + time.asctime( time.localtime(time.time()) )+ "] Charges of " + Product_Cust_ID + " is not matching \n")                                            Log_File.write("\n Baseline Value:\n")                                            Log_File.writelines( Lines_Base[0] + "\n" )                    Log_File.write("\n New Value:\n")                                            Log_File.writelines(Lines_New[0]+ "\n" )                                         Validation_Rows = Validation_Rows -1                                        TC_Row = TC_Row + 1                else:                    Master_Compare_sheet.cell(row =TC_Row,column = 5).value = "Matching"                    Master_workbook.save('C:/Test1/Logs/'+ now.strftime("%Y-%m-%d-%H-%M-%S")+ '/Master_File_Results.xlsx')                                        Validation_Rows = Validation_Rows -1                    TC_Row = TC_Row + 1                                elif (Baseline_PCID_Count == 0 and FilesToCompare_PCID_Count != 0 ):                Master_Compare_sheet.cell(row =TC_Row,column = 5).value = "PCID not found in baseline file"                Master_workbook.save('C:/Test1/Logs/'+ now.strftime("%Y-%m-%d-%H-%M-%S")+ '/Master_File_Results.xlsx')                                Validation_Rows = Validation_Rows -1                TC_Row = TC_Row + 1            elif (Baseline_PCID_Count != 0 and FilesToCompare_PCID_Count == 0 ):                Master_Compare_sheet.cell(row =TC_Row,column = 5).value = "PCID not found in new file"                Master_workbook.save('C:/Test1/Logs/'+ now.strftime("%Y-%m-%d-%H-%M-%S")+ '/Master_File_Results.xlsx')                                Validation_Rows = Validation_Rows -1                                TC_Row = TC_Row + 1            else:                Master_Compare_sheet.cell(row =TC_Row,column = 5).value = "PCID not found in both files"                Master_workbook.save('C:/Test1/Logs/'+ now.strftime("%Y-%m-%d-%H-%M-%S")+ '/Master_File_Results.xlsx')                Validation_Rows = Validation_Rows -1                TC_Row = TC_Row + 1    else:        Inc = 0                for Baseline_Row in range(0, Baseline_PCID_Count):                        status = "True"                        Acc_Found = "True"                        for Newfile_Row in range(0, FilesToCompare_PCID_Count):                Base_Acc_No = Lines_Base[Baseline_Row]                                New_Acc_No = Lines_New[Newfile_Row]                                base_Acc_No = Base_Acc_No [36:48].strip()                                new_Acc_No = New_Acc_No [36:48].strip()                if base_Acc_No == new_Acc_No:                                        Acc_Found = "Test"                                    if Lines_Base[Baseline_Row] == Lines_New[Newfile_Row]:                                            if status == "True":                                                        Master_Compare_sheet.cell(row =TC_Row,column = 5).value = "Multiple Entries- Matching"                            Master_workbook.save("C:/Test1/Logs/" + now.strftime("%Y-%m-%d-%H-%M-%S") + "/Master_File_Results.xlsx")                                                break                    else:                        status = "False"                                                Master_Compare_sheet.cell(row =TC_Row,column = 5).value = "Multiple Entries- Not Matching"                        Log_File.write("\n")                        Log_File.writelines("\n[" + time.asctime( time.localtime(time.time()) )+ "] Charges of " + Product_Cust_ID + " is not matching \n")                                                Log_File.write("Baseline Value:\n")                                                    Log_File.writelines( Lines_Base[Baseline_Row] + "\n" )                        Log_File.write("\nNew Value:\n")                                                    Log_File.writelines(Lines_New[Newfile_Row]+ "\n" )                        Master_Compare_sheet.cell(row =TC_Row,column = 5).fill = redFill                                                    Baseline_Value_Issue = Lines_Base[Baseline_Row]                                                New_Value_issue = Lines_New[Newfile_Row]                        Master_Compare_sheet.cell(row =TC_Row,column = 6 + Inc).value = Baseline_Value_Issue                        Master_Compare_sheet.cell(row =TC_Row,column = 7 + Inc).value = New_Value_issue                        Inc +=2                        Master_workbook.save('C:/Test1/Logs/'+ now.strftime("%Y-%m-%d-%H-%M-%S")+ '/Master_File_Results.xlsx')                                if Acc_Found != "Test":                Master_Compare_sheet.cell(row =TC_Row,column = 5).value = "Multiple Entries- Not Matching"                Log_File.writelines("\n[" + time.asctime( time.localtime(time.time()) )+ "]" " No charges for Charge Account " + str(base_Acc_No) + " in new file for PCID - " + Product_Cust_ID + "\n")                Log_File.write("\n")                                Master_Compare_sheet.cell(row =TC_Row,column = 5).fill = redFill                                Master_Compare_sheet.cell(row =TC_Row,column = 7 + Inc + 1).value = "No charges for Charge Account " + str(base_Acc_No) + " in new file \n"                                Master_workbook.save('C:/Test1/Logs/'+ now.strftime("%Y-%m-%d-%H-%M-%S")+ '/Master_File_Results.xlsx')                                                    Validation_Rows = Validation_Rows -1        TC_Row = TC_Row + 1    del Lines_New[:]        del Lines_Base[:]Log_File.close()print ( "Poster files compare completed. Please check Logs folder for results!")